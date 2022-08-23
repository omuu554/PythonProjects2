from openpyxl import load_workbook
import json

Wb = load_workbook("TestData.xlsx",read_only=False)
Sheet = Wb.active

Dic = {}

for cols in Sheet.iter_cols(min_row=1,max_row=5,min_col=2,max_col=8,values_only=True):
    TestObj = cols[0]
    TestValue = {Sheet["A2"].value : cols[1],Sheet["A3"].value : cols[2],Sheet["A4"].value : cols[3],Sheet["A5"].value : cols[4]}

    Dic[TestObj] = TestValue



print(json.dumps(Dic,indent = 4 ))
