import datetime
from openpyxl import load_workbook
from ExcelClasses import Product,Review
from RowsExcel import ProductId,ProductParent,ProductTitle,ProductCategory,ReviewId,Stars,BodyReview,DateReview,HeadlineReview,CustomerId


workbook = load_workbook(filename= "reviews-sample.xlsx",read_only=True)
Sheet = workbook.active

Products = []
Reviews = []

for rows in Sheet.iter_rows(min_row = 2 ,values_only = True):
    product = Product(id=rows[ProductId],parent=rows[ProductParent],title=rows[ProductTitle],category=rows[ProductCategory])
    Products.append(product)

    spread_date = rows[DateReview]
    parsedDate = datetime.datetime.strptime(spread_date, "%Y-%m-%d")

    review = Review(id=rows[ReviewId],customer_id=rows[CustomerId],stars=rows[Stars],headline=rows[HeadlineReview],body=rows[BodyReview],date=parsedDate)
    Reviews.append(review)


print(Products[0])
print(Reviews[0])





