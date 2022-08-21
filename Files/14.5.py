from openpyxl import Workbook


Wb = Workbook()
Sheet = Wb.active

Num = 1


for cols in range(2,9):
   Cell =  Sheet.cell(row=1,column=cols)
   Cell.value = f"Test {Num}"
   Num += 1

Sheet["A2"] = "Category"
Sheet["A3"] = "ProductId"
Sheet["A4"] = "Quantity"
Sheet["A5"] = "Color"

for rows in range(2,6):
   for cols in range(2,9):
      Cell = Sheet.cell(row=rows, column=cols)
      Cell.value = input(f"Please enter a value into cell row {rows} in col {cols}: ")
      if (Cell.value.isdigit()):
         Cell.value = int(Cell.value)




Wb.save("TestData.xlsx")



