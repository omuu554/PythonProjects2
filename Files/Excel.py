from openpyxl import Workbook
from openpyxl import load_workbook
import json

def CreateJsonFile(sheet2):
    ProductDic = {}
    for rows in sheet2.iter_rows(min_row= 2,min_col= 4,max_col= 7,values_only=True):
       Pruduct_id =  rows[0]
       Product = {

         "parent": rows[1],
         "title": rows[2],
         "category": rows[3]
       }
       ProductDic[Pruduct_id] = Product

    return ProductDic


workbook2 = load_workbook(filename= "reviews-sample.xlsx")
sheet2 = workbook2.active



Dic = CreateJsonFile(sheet2)
print(json.dumps(Dic))







