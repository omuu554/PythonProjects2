from openpyxl import Workbook
from openpyxl import load_workbook

"""Wb = Workbook()
Sheet = Wb.active

Sheet["A1"] = "UserName"
Sheet["B1"] = "Password"
Sheet["A2"] = "Tester1"
Sheet["B2"] = "1234"
Sheet["A3"] = "Tester2"
Sheet["B3"] = "45678"

Wb.save("Login.xlsx")"""

WorkingXl = load_workbook("Login.xlsx",read_only=False)
Sheet = WorkingXl.active

print(*WorkingXl.sheetnames)

"""for Rows in Sheet.iter_rows(min_row=2,max_row=3,values_only=True):
    TupleRow = Rows
    for i in range(len(TupleRow)):
        print(TupleRow[i])"""

for Row in range(2,4):
    for Col in range(1,3):
        print(Sheet.cell(row=Row,column=Col).value)


for Rows in Sheet.iter_rows(min_row=1,max_row=3,values_only=True):
    TupleRow = Rows
    for i in range(len(TupleRow)):
        print(TupleRow[i])


for Cols in Sheet.iter_cols(min_col=1,max_col=2,values_only=True):
    TupleCol = Cols
    for i in range(len(TupleCol)):
        print(TupleCol[i])

